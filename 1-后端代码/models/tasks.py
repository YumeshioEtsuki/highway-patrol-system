# models/tasks.py
import os
import time
from datetime import datetime
import mysql.connector
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from io import BytesIO
from fastapi.responses import StreamingResponse
from PIL import Image, ImageDraw, ImageFont
from utils.config import db_config, settings
from utils.test_data import get_test_data
from utils.sse import sse_message
from utils.algorithm import hash_password, verify_password
from utils.utils import get_db_connection
from routes.patrol_sse import push_new_photo_event
from .schema import CREATE_TABLES_SQL

# 统计缓存（优先使用 Redis，失败则使用内存）
_STATS_CACHE = {}
_STATS_DEFAULT_TTL = int(os.getenv("STATS_CACHE_TTL", "600"))  # 秒
try:
    import redis
    _redis_client = redis.from_url(os.getenv("REDIS_URL", "")) if os.getenv("REDIS_URL") else None
except Exception:
    _redis_client = None

def _cache_get(key: str):
    if _redis_client:
        try:
            val = _redis_client.get(key)
            if val:
                import json as _json
                return _json.loads(val)
        except Exception:
            pass
        return None
    item = _STATS_CACHE.get(key)
    if not item:
        return None
    if item['expires_at'] < time.time():
        _STATS_CACHE.pop(key, None)
        return None
    return item['value']

def _cache_set(key: str, value, ttl: int | None = None):
    ttl = ttl or _STATS_DEFAULT_TTL
    if _redis_client:
        try:
            import json as _json
            _redis_client.setex(key, ttl, _json.dumps(value, ensure_ascii=False))
            return
        except Exception:
            pass
    _STATS_CACHE[key] = {
        'value': value,
        'expires_at': time.time() + ttl
    }

def _cache_clear(pattern: str = "admin_stats:"):
    """清除指定前缀的缓存"""
    if _redis_client:
        try:
            # Redis 模式清理
            cursor = 0
            while True:
                cursor, keys = _redis_client.scan(cursor, match=f"{pattern}*")
                if keys:
                    _redis_client.delete(*keys)
                if cursor == 0:
                    break
            return
        except Exception:
            pass
    # 内存缓存清理
    for key in list(_STATS_CACHE.keys()):
        if key.startswith(pattern):
            _STATS_CACHE.pop(key, None)


os.makedirs(settings.UPLOAD_FOLDER, exist_ok=True)


def insert_audit_log(user_id: int, action: str, resource: str, details: str | None = None):
    """写入审计日志（若表不存在则忽略错误）"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO AuditLog (user_id, action, resource, details)
            VALUES (%s, %s, %s, %s)
            """,
            (user_id, action, resource, details)
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception:
        # 容忍审计失败，不影响主流程
        try:
            cur.close()
            conn.close()
        except Exception:
            pass


def get_audit_logs(action: str | None = None, user_id: int | None = None,
                   start_date: str | None = None, end_date: str | None = None,
                   keyword: str | None = None, page: int = 1, page_size: int = 50):
    """查询审计日志（支持筛选与分页）"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        sql = """
            SELECT id, user_id, action, resource, details, timestamp
            FROM AuditLog
            WHERE 1=1
        """
        params = []

        count_sql = """
            SELECT COUNT(*) AS total
            FROM AuditLog
            WHERE 1=1
        """
        count_params = []

        if action:
            sql += " AND action = %s"
            params.append(action)
            count_sql += " AND action = %s"
            count_params.append(action)
        if user_id:
            sql += " AND user_id = %s"
            params.append(user_id)
            count_sql += " AND user_id = %s"
            count_params.append(user_id)
        if start_date:
            sql += " AND DATE(timestamp) >= %s"
            params.append(start_date)
            count_sql += " AND DATE(timestamp) >= %s"
            count_params.append(start_date)
        if end_date:
            sql += " AND DATE(timestamp) <= %s"
            params.append(end_date)
            count_sql += " AND DATE(timestamp) <= %s"
            count_params.append(end_date)
        if keyword:
            sql += " AND (resource LIKE %s OR details LIKE %s)"
            like_kw = f"%{keyword}%"
            params.extend([like_kw, like_kw])
            count_sql += " AND (resource LIKE %s OR details LIKE %s)"
            count_params.extend([like_kw, like_kw])

        # 分页约束
        try:
            page = int(page) if page and int(page) > 0 else 1
        except Exception:
            page = 1
        try:
            page_size = int(page_size) if page_size else 50
        except Exception:
            page_size = 50
        page_size = max(1, min(page_size, settings.MAX_PAGE_SIZE))
        offset = (page - 1) * page_size

        sql += " ORDER BY timestamp DESC LIMIT %s OFFSET %s"

        cursor.execute(count_sql, count_params)
        total = cursor.fetchone()["total"]

        cursor.execute(sql, params + [page_size, offset])
        rows = cursor.fetchall()
        return {
            'records': rows,
            'total': int(total or 0),
            'page': page,
            'page_size': page_size
        }
    finally:
        cursor.close()
        conn.close()


# ========================
# 用户相关
# ========================

def register_user(username, password, real_name, phone, email, role):
    # 🔑 关键：先对明文密码进行哈希
    hashed_pwd = hash_password(password)

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO User (username, password, real_name, phone, email, role)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (username, hashed_pwd, real_name, phone, email, role))  # 👈 存的是 hashed_pwd
        conn.commit()
        return cursor.lastrowid
    except Exception as e:
        conn.rollback()
        if "Duplicate entry" in str(e):
            raise ValueError("用户名或手机号已存在")
        else:
            raise e
    finally:
        cursor.close()
        conn.close()


def user_login_by_password(username, password):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM User WHERE username = %s", (username,))
        user = cursor.fetchone()

        if not user:
            return None

        stored_pwd = user['password']

        # 兼容旧数据：明文、旧版 SHA256(hex) 也能迁移
        password_ok = verify_password(stored_pwd, password)

        # 1) 明文存储
        if not password_ok and stored_pwd == password:
            new_hash = hash_password(password)
            cursor.execute("UPDATE User SET password = %s WHERE user_id = %s", (new_hash, user['user_id']))
            conn.commit()
            password_ok = True

        # 2) 旧版无盐 SHA256(hex) 兼容
        if not password_ok and len(stored_pwd) == 64:
            import hashlib
            if stored_pwd.lower() == hashlib.sha256(password.encode()).hexdigest():
                new_hash = hash_password(password)
                cursor.execute("UPDATE User SET password = %s WHERE user_id = %s", (new_hash, user['user_id']))
                conn.commit()
                password_ok = True

        if password_ok:
            cursor.execute("UPDATE User SET last_login = NOW() WHERE user_id = %s", (user['user_id'],))
            conn.commit()
            del user['password']
            return user

        return None
    finally:
        cursor.close()
        conn.close()


def update_user_password(user_id, old_password, new_password):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # 1. 查询当前密码哈希
        cursor.execute("SELECT password FROM User WHERE user_id = %s", (user_id,))
        row = cursor.fetchone()

        # 2. 验证旧密码是否正确
        if not row or not verify_password(row[0], old_password):  # ← 使用封装的验证
            return False

        # 3. 生成新密码的哈希
        new_hash = hash_password(new_password)  # ← 使用封装的哈希

        # 4. 更新数据库
        cursor.execute("UPDATE User SET password = %s WHERE user_id = %s", (new_hash, user_id))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cursor.close()
        conn.close()



# ========================
# 路段和问题类型相关
# ========================

def get_all_road_segments():
    """获取所有路段信息"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT segment_id, segment_name, start_number, end_number, department_id, region
            FROM RoadSegment
            ORDER BY segment_id
        """)
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()


def get_all_problem_types():
    """获取所有问题类型"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT type_id, type_name, parent_id
            FROM ProblemType
            ORDER BY type_id
        """)
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()


# ========================
# 巡查记录相关
# ========================

def create_patrol_record(data):
    required = ['user_id', 'patrol_time', 'segment_id', 'problem_type_id', 'description']
    for field in required:
        if field not in data:
            raise ValueError(f"缺少必要字段: {field}")

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO InspectionRecord 
            (user_id, upload_time, segment_id, problem_type_id, description, severity, latitude, longitude, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data['user_id'],
            data['patrol_time'],
            data['segment_id'],
            data['problem_type_id'],
            data['description'],
            data.get('severity', 1),
            data.get('latitude'),
            data.get('longitude'),
            'pending'
        ))
        conn.commit()
        return cursor.lastrowid
    finally:
        cursor.close()
        conn.close()


def get_patrol_list(user_id, page=1, page_size=10):
    try:
        page = int(page) if page and int(page) > 0 else 1
    except Exception:
        page = 1
    try:
        page_size = int(page_size) if page_size else 10
    except Exception:
        page_size = 10
    page_size = max(1, min(page_size, settings.MAX_PAGE_SIZE))
    offset = (page - 1) * page_size
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT COUNT(*) AS total FROM InspectionRecord WHERE user_id = %s", (user_id,))
        total = cursor.fetchone()['total']

        cursor.execute("""
            SELECT ir.*, 
                   rs.segment_name,
                   pt.type_name AS problem_type_name,
                   u.real_name AS inspector_name
            FROM InspectionRecord ir
            LEFT JOIN RoadSegment rs ON ir.segment_id = rs.segment_id
            LEFT JOIN ProblemType pt ON ir.problem_type_id = pt.type_id
            LEFT JOIN User u ON ir.user_id = u.user_id
            WHERE ir.user_id = %s
            ORDER BY ir.upload_time DESC
            LIMIT %s OFFSET %s
        """, (user_id, page_size, offset))

        records = cursor.fetchall()
        return {
            'records': records,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    finally:
        cursor.close()
        conn.close()


def get_patrol_detail(record_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT ir.*, 
                   rs.segment_name,
                   pt.type_name AS problem_type_name,
                   u.real_name AS inspector_name,
                   d.department_name
            FROM InspectionRecord ir
            LEFT JOIN RoadSegment rs ON ir.segment_id = rs.segment_id
            LEFT JOIN ProblemType pt ON ir.problem_type_id = pt.type_id
            LEFT JOIN User u ON ir.user_id = u.user_id
            LEFT JOIN Department d ON u.department_id = d.department_id
            WHERE ir.record_id = %s
        """, (record_id,))
        record = cursor.fetchone()

        if not record:
            return None

        cursor.execute("SELECT * FROM Photo WHERE record_id = %s", (record_id,))
        photos = cursor.fetchall()
        record['photos'] = photos
        return record
    finally:
        cursor.close()
        conn.close()



# ========================
# 照片上传
# ========================

def save_photo_to_db(record_id, file_path, file_name):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO Photo (record_id, photo_type, file_path, file_name)
            VALUES (%s, 'test_pictures', %s, %s)
        """, (record_id, file_path, file_name))
        conn.commit()
        return cursor.lastrowid
    finally:
        cursor.close()
        conn.close()


def add_watermark(image_path, gps_info="(39.915, 116.404)", address="北京市东城区", timestamp=None):
    """给照片添加水印"""
    if not timestamp:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    img = Image.open(image_path).convert("RGBA")
    txt = Image.new('RGBA', img.size, (255, 255, 255, 0))

    # 使用默认字体（或指定字体文件）
    try:
        font = ImageFont.truetype("simhei.ttf", 30)  # 中文字体
    except:
        font = ImageFont.load_default()

    d = ImageDraw.Draw(txt)
    watermark_text = f"{timestamp}\n{gps_info}\n{address}"

    # 放在左下角
    d.text((10, img.height - 100), watermark_text, fill=(255, 255, 255, 200), font=font)

    watermarked = Image.alpha_composite(img, txt)
    watermarked.convert("RGB").save(image_path, "JPEG")



# ========================
# 统计
# ========================

def get_user_stats(user_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT COUNT(*) AS total FROM InspectionRecord WHERE user_id = %s", (user_id,))
        total = cursor.fetchone()['total']

        cursor.execute("""
            SELECT status, COUNT(*) AS count
            FROM InspectionRecord
            WHERE user_id = %s
            GROUP BY status
        """, (user_id,))
        status_counts = {row['status']: row['count'] for row in cursor.fetchall()}

        cursor.execute("""
            SELECT COUNT(*) AS recent
            FROM InspectionRecord
            WHERE user_id = %s AND upload_time >= DATE_SUB(NOW(), INTERVAL 7 DAY)
        """, (user_id,))
        recent = cursor.fetchone()['recent']

        return {
            'total_records': total,
            'status_breakdown': status_counts,
            'recent_7_days': recent
        }
    finally:
        cursor.close()
        conn.close()


def get_admin_stats(start_date=None, end_date=None, region=None, scope=None, province=None, city=None,
                   lat_min=None, lat_max=None, lon_min=None, lon_max=None, data_type=None):
    """管理员统计：总数、状态分布、问题类型分布、严重度分布、近7/30天新增

    Args:
        start_date: 开始日期
        end_date: 结束日期
        region: 区域过滤（大洲名称），如: 'Asia', 'Europe', 'North America'等 [已过时，兼容保留]
        scope: 数据范围 'world' | 'china' | 'province' | 'city'
        province: 省份名称（如'浙江省'、'北京市'）
        city: 城市名称（需配合province使用）
        data_type: 数据类型筛选 'real' | 'test' | None(全部)
    
    Returns:
        包含统计数据的字典
    """
    from models.china_regions import get_province_gps_bounds, get_city_gps_bounds, CHINA_PROVINCES_GPS

    # 构建缓存键（不同筛选条件独立缓存）
    cache_key = f"admin_stats:{start_date}:{end_date}:{region}:{scope}:{province}:{city}:{lat_min}:{lat_max}:{lon_min}:{lon_max}:{data_type}"
    cached = _cache_get(cache_key)
    if cached:
        return cached

    def normalize_province_name(name: str | None) -> str | None:
        """容错省份名称（去空格、补/去后缀）"""
        if not name:
            return None
        raw = name.strip()
        if raw in CHINA_PROVINCES_GPS:
            return raw

        # 去除常见后缀后尝试匹配
        def strip_suffix(val: str) -> str:
            return val.replace("省", "").replace("市", "").replace("自治区", "").replace("特别行政区", "")

        simplified = strip_suffix(raw)
        for prov in CHINA_PROVINCES_GPS.keys():
            if simplified == strip_suffix(prov):
                return prov
        return raw

    def normalize_city_name(province_name: str | None, city_name: str | None) -> str | None:
        """容错城市名称，兼容“杭州市”/“杭州”等写法"""
        if not province_name or not city_name:
            return city_name.strip() if city_name else None

        cities = CHINA_PROVINCES_GPS.get(province_name, {}).get("cities", {})
        if not cities:
            return city_name.strip()

        def strip_suffix(val: str) -> str:
            return val.replace("市", "").replace("区", "").replace("县", "").replace("盟", "").replace("自治州", "").strip()

        simplified_target = strip_suffix(city_name)
        for cname in cities.keys():
            if simplified_target == strip_suffix(cname):
                return cname
        return city_name.strip()

    normalized_province = normalize_province_name(province)
    normalized_city = normalize_city_name(normalized_province, city)

    # 大洲粗粒度边界（用于世界视图区域过滤）
    region_bounds = {
        'Asia': {'lat_min': -15, 'lat_max': 60, 'lon_min': 25, 'lon_max': 150},
        'Europe': {'lat_min': 34, 'lat_max': 72, 'lon_min': -25, 'lon_max': 45},
        'North America': {'lat_min': 5, 'lat_max': 83, 'lon_min': -170, 'lon_max': -50},
        'South America': {'lat_min': -60, 'lat_max': 15, 'lon_min': -85, 'lon_max': -30},
        'Africa': {'lat_min': -35, 'lat_max': 38, 'lon_min': -20, 'lon_max': 55},
        'Oceania': {'lat_min': -50, 'lat_max': 10, 'lon_min': 110, 'lon_max': 180}
    }
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        where_clause = " WHERE 1=1"
        params = []
        
        if start_date:
            where_clause += " AND DATE(ir.upload_time) >= %s"
            params.append(start_date)
        if end_date:
            where_clause += " AND DATE(ir.upload_time) <= %s"
            params.append(end_date)
        
        # 数据类型筛选
        if data_type in ['real', 'test']:
            where_clause += " AND ir.data_type = %s"
            params.append(data_type)
        
        # 优先使用自定义经纬度边界
        custom_bounds = all(v is not None for v in [lat_min, lat_max, lon_min, lon_max])
        if custom_bounds:
            where_clause += " AND ir.latitude BETWEEN %s AND %s"
            where_clause += " AND ir.longitude BETWEEN %s AND %s"
            params.extend([lat_min, lat_max, lon_min, lon_max])
        else:
            # GPS过滤：根据scope、province、city确定地理边界
            if scope == 'china':
                # 中国大陆大致边界框
                where_clause += " AND ir.latitude BETWEEN 18 AND 54"
                where_clause += " AND ir.longitude BETWEEN 73 AND 135"
            elif scope == 'province' and normalized_province:
                bounds = get_province_gps_bounds(normalized_province)
                if bounds:
                    where_clause += f" AND ir.latitude BETWEEN {bounds['lat_min']} AND {bounds['lat_max']}"
                    where_clause += f" AND ir.longitude BETWEEN {bounds['lon_min']} AND {bounds['lon_max']}"
            elif scope == 'city' and normalized_province and normalized_city:
                bounds = get_city_gps_bounds(normalized_province, normalized_city)
                if bounds:
                    where_clause += f" AND ir.latitude BETWEEN {bounds['lat_min']} AND {bounds['lat_max']}"
                    where_clause += f" AND ir.longitude BETWEEN {bounds['lon_min']} AND {bounds['lon_max']}"
            elif scope == 'world' and region and region_bounds.get(region):
                rb = region_bounds[region]
                where_clause += f" AND ir.latitude BETWEEN {rb['lat_min']} AND {rb['lat_max']}"
                where_clause += f" AND ir.longitude BETWEEN {rb['lon_min']} AND {rb['lon_max']}"

        # 总数
        cursor.execute(f"SELECT COUNT(*) AS total FROM InspectionRecord ir{where_clause}", params)
        total = cursor.fetchone()['total']

        # 状态分布
        cursor.execute(f"""
            SELECT ir.status, COUNT(*) AS count
            FROM InspectionRecord ir
            {where_clause}
            GROUP BY ir.status
        """, params)
        status_breakdown = {row['status']: row['count'] for row in cursor.fetchall()}

        # 问题类型分布
        cursor.execute(f"""
            SELECT COALESCE(pt.type_name, '未分类') AS label, COUNT(*) AS count
            FROM InspectionRecord ir
            LEFT JOIN ProblemType pt ON ir.problem_type_id = pt.type_id
            {where_clause}
            GROUP BY label
            ORDER BY count DESC
            LIMIT 10
        """, params)
        type_breakdown = cursor.fetchall()

        # 严重度分布
        cursor.execute(f"""
            SELECT ir.severity AS label, COUNT(*) AS count
            FROM InspectionRecord ir
            {where_clause}
            GROUP BY ir.severity
            ORDER BY ir.severity
        """, params)
        severity_breakdown = cursor.fetchall()

        # 近7天、30天
        cursor.execute(f"""
            SELECT COUNT(*) AS cnt
            FROM InspectionRecord ir
            {where_clause} AND ir.upload_time >= DATE_SUB(NOW(), INTERVAL 7 DAY)
        """, params)
        recent_7_days = cursor.fetchone()['cnt']

        cursor.execute(f"""
            SELECT COUNT(*) AS cnt
            FROM InspectionRecord ir
            {where_clause} AND ir.upload_time >= DATE_SUB(NOW(), INTERVAL 30 DAY)
        """, params)
        recent_30_days = cursor.fetchone()['cnt']

        result = {
            'total_records': total,
            'status_breakdown': status_breakdown,
            'type_breakdown': type_breakdown,
            'severity_breakdown': severity_breakdown,
            'recent_7_days': recent_7_days,
            'recent_30_days': recent_30_days,
            'scope': scope or 'world',
            'region': normalized_city or normalized_province or region or 'global'
        }
        _cache_set(cache_key, result)
        return result
    finally:
        cursor.close()
        conn.close()



# ========================
# 管理员操作：状态流转、获取所有巡查记录
# ========================

def mark_record_as_processing(record_id: int):
    """
    将巡查记录状态从 'pending' 改为 'processing'
    返回是否成功（布尔值）
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # 只允许从 pending → processing
        cursor.execute("""
            UPDATE InspectionRecord 
            SET status = 'processing', admin_process_time = %s
            WHERE record_id = %s AND status = 'pending'
        """, (datetime.now(), record_id))

        success = cursor.rowcount > 0
        conn.commit()
        return success
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cursor.close()
        conn.close()


def mark_record_as_completed(record_id: int, process_note: str = "", fix_time=None):
    """
    将巡查记录状态从 'processing' 改为 'completed'
    必须提供修复备注（可为空字符串），修复时间默认为当前时间
    """
    if fix_time is None:
        fix_time = datetime.now()

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # 只允许从 processing → completed
        cursor.execute("""
            UPDATE InspectionRecord 
            SET 
                status = 'completed',
                fix_time = %s,
                process_note = %s
            WHERE record_id = %s AND status = 'processing'
        """, (fix_time, process_note, record_id))

        success = cursor.rowcount > 0
        conn.commit()
        return success
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cursor.close()
        conn.close()


def get_patrol_list_admin(status_filter=None, problem_type_id=None, severity=None,
                          start_date=None, end_date=None, keyword=None, data_type=None,
                          page: int = 1, page_size: int = 100):
    """
    获取所有巡查记录（供管理员使用）
    支持：状态 / 问题类型 / 严重度 / 起止日期 / 关键词（描述、路段、问题类型）筛选
    """
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        sql = """
            SELECT 
                ir.record_id,
                ir.user_id,
                ir.segment_id,
                ir.problem_type_id,
                ir.description,
                ir.severity,
                ir.status,
                ir.upload_time,
                ir.latitude,
                ir.longitude,
                ir.data_type,
                rs.segment_name,
                pt.type_name AS problem_type,
                u.real_name AS reporter
            FROM InspectionRecord ir
            LEFT JOIN RoadSegment rs ON ir.segment_id = rs.segment_id
            LEFT JOIN ProblemType pt ON ir.problem_type_id = pt.type_id
            LEFT JOIN User u ON ir.user_id = u.user_id
            WHERE 1=1
        """
        params = []

        # 统计总数SQL（与上面过滤条件保持一致）
        count_sql = """
            SELECT COUNT(*) AS total
            FROM InspectionRecord ir
            LEFT JOIN RoadSegment rs ON ir.segment_id = rs.segment_id
            LEFT JOIN ProblemType pt ON ir.problem_type_id = pt.type_id
            LEFT JOIN User u ON ir.user_id = u.user_id
            WHERE 1=1
        """
        count_params = []

        if status_filter:
            sql += " AND ir.status = %s"
            params.append(status_filter)
            count_sql += " AND ir.status = %s"
            count_params.append(status_filter)
        if problem_type_id:
            sql += " AND ir.problem_type_id = %s"
            params.append(problem_type_id)
            count_sql += " AND ir.problem_type_id = %s"
            count_params.append(problem_type_id)

        if severity:
            sql += " AND ir.severity = %s"
            params.append(severity)
            count_sql += " AND ir.severity = %s"
            count_params.append(severity)
        
        if data_type in ['real', 'test']:
            sql += " AND ir.data_type = %s"
            params.append(data_type)
            count_sql += " AND ir.data_type = %s"
            count_params.append(data_type)

        if start_date:
            sql += " AND DATE(ir.upload_time) >= %s"
            params.append(start_date)
            count_sql += " AND DATE(ir.upload_time) >= %s"
            count_params.append(start_date)

        if end_date:
            sql += " AND DATE(ir.upload_time) <= %s"
            params.append(end_date)
            count_sql += " AND DATE(ir.upload_time) <= %s"
            count_params.append(end_date)

        if keyword:
            sql += " AND (ir.description LIKE %s OR rs.segment_name LIKE %s OR pt.type_name LIKE %s)"
            like_kw = f"%{keyword}%"
            params.extend([like_kw, like_kw, like_kw])
            count_sql += " AND (ir.description LIKE %s OR rs.segment_name LIKE %s OR pt.type_name LIKE %s)"
            count_params.extend([like_kw, like_kw, like_kw])

        # 分页（限制单页最大 settings.MAX_PAGE_SIZE，防止一次性返回过多导致前端卡顿）
        try:
            page = int(page) if page and int(page) > 0 else 1
        except Exception:
            page = 1
        try:
            page_size = int(page_size) if page_size else 100
        except Exception:
            page_size = 100
        page_size = max(1, min(page_size, settings.MAX_PAGE_SIZE))
        offset = (page - 1) * page_size

        sql += " ORDER BY ir.upload_time DESC LIMIT %s OFFSET %s"

        # 查询总数
        cursor.execute(count_sql, count_params)
        total = cursor.fetchone()["total"]

        # 查询分页数据
        cursor.execute(sql, params + [page_size, offset])
        records = cursor.fetchall()
        return {
            'records': records,
            'total': int(total or 0),
            'page': page,
            'page_size': page_size
        }
    finally:
        cursor.close()
        conn.close()



# ========================
# 数据导出
# ========================

def export_patrol_records_to_excel(filters: dict = None) -> bytes:
    """
    根据筛选条件导出巡查记录为 Excel 文件（bytes）
    支持自动列宽、中文列名、时间格式化等优化
    """
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    base_sql = """
        SELECT 
            r.record_id,
            u.real_name AS reporter,
            s.segment_name,
            pt.type_name AS problem_type,
            r.description,
            r.severity,
            r.status,
            r.upload_time,
            r.admin_process_time,
            r.fix_time,
            r.process_note
        FROM InspectionRecord r
        JOIN User u ON r.user_id = u.user_id
        JOIN RoadSegment s ON r.segment_id = s.segment_id
        JOIN ProblemType pt ON r.problem_type_id = pt.type_id
        WHERE 1=1
    """
    params = []

    if filters:
        if filters.get('start_date'):
            base_sql += " AND r.upload_time >= %s"
            params.append(filters['start_date'])
        if filters.get('end_date'):
            base_sql += " AND r.upload_time <= %s"
            params.append(filters['end_date'] + ' 23:59:59')
        if filters.get('segment_id'):
            base_sql += " AND r.segment_id = %s"
            params.append(filters['segment_id'])
        if filters.get('status'):
            base_sql += " AND r.status = %s"
            params.append(filters['status'])
        if filters.get('problem_type_id'):
            base_sql += " AND r.problem_type_id = %s"
            params.append(filters['problem_type_id'])
        if filters.get('severity'):
            base_sql += " AND r.severity = %s"
            params.append(filters['severity'])
        if filters.get('keyword'):
            base_sql += " AND (r.description LIKE %s OR pt.type_name LIKE %s OR s.segment_name LIKE %s OR u.real_name LIKE %s)"
            kw_like = f"%{filters['keyword']}%"
            params.extend([kw_like, kw_like, kw_like, kw_like])

    base_sql += " ORDER BY r.upload_time DESC"

    try:
        cursor.execute(base_sql, params)
        records = cursor.fetchall()

        # 即使无数据，也保留表结构（方便用户知道有哪些字段）
        if not records:
            records = [{
                'record_id': None,
                'reporter': None,
                'segment_name': None,
                'problem_type': None,
                'description': None,
                'severity': None,
                'status': None,
                'upload_time': None,
                'admin_process_time': None,
                'fix_time': None,
                'process_note': None
            }]

        df = pd.DataFrame(records)

        # 中文列名映射
        col_names = {
            'record_id': '记录ID',
            'reporter': '上报人',
            'segment_name': '路段名称',
            'problem_type': '问题类型',
            'description': '问题描述',
            'severity': '严重程度',
            'status': '状态',
            'upload_time': '上报时间',
            'admin_process_time': '管理员处理时间',
            'fix_time': '修复完成时间',
            'process_note': '处理备注'
        }
        df.rename(columns=col_names, inplace=True)

        # 状态值本地化
        status_map = {'pending': '待处理', 'processing': '处理中', 'resolved': '已修复'}
        df['状态'] = df['状态'].map(status_map).fillna(df['状态'])

        # 时间列格式化为字符串（避免 Excel 显示为数字）
        time_cols = ['上报时间', '管理员处理时间', '修复完成时间']
        for col in time_cols:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')

        # 导出到 Excel（openpyxl 引擎默认就是 UTF-8）
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='巡查记录')
            worksheet = writer.sheets['巡查记录']

            # === 自动列宽（含列名）===
            for idx, col in enumerate(df.columns, 1):
                header_len = len(str(col))
                if df.empty:
                    content_max = 0
                else:
                    content_max = df[col].fillna('').astype(str).apply(len).max()
                width = min(max(header_len, content_max) + 2, 50)
                worksheet.column_dimensions[get_column_letter(idx)].width = width

                # 设置换行（可选）
                for cell in worksheet[get_column_letter(idx)][1:]:  # 从第2行开始（跳过表头）
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

        output.seek(0)
        return output.getvalue()

    finally:
        cursor.close()
        conn.close()




# ========================
# 图片获取
# ========================

def get_photo_by_id(photo_id):
    """从数据库获取图片二进制数据"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT photo_data FROM photo WHERE photo_id = %s", (photo_id,))
        row = cursor.fetchone()
        if row:
            return row['photo_data']  # bytes
        else:
            return None
    except Exception as e:
        print(f"获取图片失败: {e}")
        return None
    finally:
        if conn:
            conn.close()



# ========================
# 管理员页面
# ========================

def stream_verify_database():
    def generate():
        try:
            yield sse_message('log', '开始验证数据库完整性...')
            time.sleep(0.05)

            conn = None
            cursor = None
            try:
                conn = mysql.connector.connect(connect_timeout=3, **db_config)
                cursor = conn.cursor()
                cursor.execute("SHOW TABLES LIKE 'inspectionrecord'")
                exists = len(cursor.fetchall()) > 0
                if exists:
                    yield sse_message('success', '✅ 表 inspectionrecord 存在', step=1)
                else:
                    yield sse_message('error', '❌ 表 inspectionrecord 不存在', step=1)

                yield sse_message('log', '🔍 正在验证数据可读性...', step=2)
                time.sleep(0.05)
                cursor.execute("SELECT record_id FROM inspectionrecord LIMIT 1")
                has_data = cursor.fetchone() is not None
                if has_data:
                    yield sse_message('success', '✅ 数据可正常读取', step=2)
                else:
                    yield sse_message('warn', '⚠️ 表存在但无数据', step=2)
            except Exception as e:
                yield sse_message('error', f'❌ 数据库连接/查询失败: {e}', step=0)
            finally:
                if cursor:
                    cursor.close()
                if conn:
                    conn.close()

            yield sse_message('complete', '🎉 数据库验证完成！')
        except Exception as e:
            yield sse_message('error', f'❌ 内部错误: {str(e)}')
    return StreamingResponse(
        generate(),
        media_type='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no'
        }
    )


# ========================
# 生成随机测试数据
# ========================

def generate_fake_records(count: int = 50, with_photos: bool = False):
    """生成测试数据，支持地理位置分布到各省份/城市
    
    args:
        count: 生成数量
        with_photos: 是否生成照片
    
    注：GPS坐标会根据随机选择的省份分布，确保数据具有地理有效性
    """
    import random
    from datetime import timedelta
    from PIL import Image, ImageDraw, ImageFont
    from models.china_regions import CHINA_PROVINCES_GPS
    
    photos_dir = settings.UPLOAD_FOLDER
    os.makedirs(photos_dir, exist_ok=True)

    conn = None
    inserted_count = 0
    failed_count = 0
    try:
        conn = mysql.connector.connect(
            host=db_config['host'],
            user=db_config['user'],
            password=db_config['password'],
            database=db_config['database'],
            charset='utf8mb4',
            autocommit=False
        )
        cursor = conn.cursor()

        # 获取用户、路段、问题类型
        cursor.execute("SELECT user_id FROM User ORDER BY user_id LIMIT 1")
        user_row = cursor.fetchone()
        if not user_row:
            raise ValueError("无可用用户，请先初始化数据库")
        user_id = user_row[0]

        cursor.execute("SELECT segment_id FROM RoadSegment")
        segments = [r[0] for r in cursor.fetchall()] or [None]

        cursor.execute("SELECT type_id FROM ProblemType")
        types = [r[0] for r in cursor.fetchall()] or [None]

        now = datetime.now()
        statuses = ['pending', 'processing', 'resolved']
        
        # 获取所有省份列表用于GPS分布
        all_provinces = list(CHINA_PROVINCES_GPS.keys())
        
        # 提前创建插入SQL语句，确保批量插入效率
        record_ids = []
        
        # 定期commit的批次大小
        BATCH_SIZE = 100

        for i in range(int(count)):
            try:
                dt = now - timedelta(days=random.randint(0, 30), seconds=random.randint(0, 86400))
                
                # 随机选择一个省份，然后在该省份内生成GPS坐标
                province = random.choice(all_provinces)
                province_data = CHINA_PROVINCES_GPS[province]
                lat_min, lat_max = province_data["lat_range"]
                lon_min, lon_max = province_data["lon_range"]
                
                lat = round(random.uniform(lat_min, lat_max), 6)
                lon = round(random.uniform(lon_min, lon_max), 6)
                
                sev = random.randint(1, 5)
                status = random.choices(statuses, weights=[6, 3, 1])[0]
                seg = random.choice(segments)
                typ = random.choice(types)
                desc = f"[{province}] 随机测试记录 #{i+1}"

                cursor.execute(
                    """
                    INSERT INTO InspectionRecord (
                        user_id, upload_time, latitude, longitude,
                        segment_id, problem_type_id, description, severity, status, data_type
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (user_id, dt.strftime('%Y-%m-%d %H:%M:%S'), lat, lon, seg, typ, desc, sev, status, 'test')
                )
                record_id = cursor.lastrowid
                record_ids.append(record_id)
                inserted_count += 1

                if with_photos:
                    img = Image.new('RGB', (320, 200), color=(30, 40, 60))
                    draw = ImageDraw.Draw(img)
                    text = f"Record #{record_id}\nSeverity {sev}\n{lat},{lon}\n{province}"
                    try:
                        font = ImageFont.load_default()
                    except Exception:
                        font = None
                    draw.text((10, 10), text, fill=(220, 240, 255), font=font)
                    filename = f"auto_{record_id}.jpg"
                    filepath = os.path.join(photos_dir, filename)
                    img.save(filepath, format='JPEG', quality=85)

                    cursor.execute(
                        """
                        INSERT INTO Photo (record_id, photo_type, file_path, file_name, file_size, is_watermarked)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        """,
                        (record_id, 'test_pictures', filepath, filename, os.path.getsize(filepath), 1)
                    )
                    photo_id = cursor.lastrowid

                    # 推送 SSE 照片事件（若前端已订阅）
                    try:
                        # 转换为HTTP可访问的URL
                        photo_http_url = f"/photos/{filename}"
                        push_new_photo_event(record_id, photo_id, photo_http_url)
                    except Exception as e:
                        print(f"[SSE] 推送照片事件失败: {e}")
                
                # 每 BATCH_SIZE 条记录提交一次，避免大事务超时
                if (i + 1) % BATCH_SIZE == 0:
                    conn.commit()
                    print(f"[INFO] 已提交 {inserted_count} 条记录...")
                    
            except Exception as e:
                # 单条记录失败，继续处理下一条
                failed_count += 1
                print(f"[WARN] 生成第{i+1}条记录失败: {e}")
                continue

        # 最后提交剩余的数据
        conn.commit()
        cursor.close()
        conn.close()

        return {
            'success': True,
            'inserted': inserted_count,
            'requested': int(count),
            'failed': failed_count,
            'photos': bool(with_photos),
            'note': f'已生成{inserted_count}/{int(count)}条数据（失败{failed_count}条，GPS坐标按省份分布）'
        }
    except Exception as e:
        import traceback
        print(f"[ERROR] 生成测试数据失败: {e}")
        print(f"[TRACE] {traceback.format_exc()}")
        try:
            if conn:
                conn.rollback()
        except:
            pass
        return {
            'success': False,
            'error': str(e)
        }


def stream_get_database_status():
    def generate():
        yield sse_message('log', '正在获取数据库状态...')
        try:
            conn = mysql.connector.connect(
                host=db_config['host'],
                user=db_config['user'],
                password=db_config['password'],
                charset='utf8mb4',
                connect_timeout=3
            )
            yield sse_message('success', '✅ 成功连接到 MySQL 服务器')
        except Exception as e:
            yield sse_message('error', f'❌ 连接失败: {str(e)}')
            return

        try:
            cursor = conn.cursor()
            cursor.execute("SELECT VERSION() AS version")
            ver = cursor.fetchone()[0]
            yield sse_message('info', f'📦 MySQL 版本: {ver}')

            db_name = db_config['database']
            cursor.execute(f"USE `{db_name}`")
            yield sse_message('info', f'📂 当前数据库: {db_name}')

            cursor.execute("SHOW TABLES")
            tables = [row[0] for row in cursor.fetchall()]
            count = len(tables)
            yield sse_message('info', f'📊 表数量: {count}')

            if tables:
                indented_tables = [" " * 12 + "•" + " " * 2 + table for table in tables]
                table_list_str = "\n".join(indented_tables)
                yield sse_message('info', f'📋 表列表:\n{table_list_str}')
            else:
                yield sse_message('info', '📋 当前数据库为空（无表）')

            cursor.close()
            conn.close()
            yield sse_message('complete', '✅ 状态获取完成')
        except Exception as e:
            yield sse_message('error', f'❌ 获取状态时出错: {str(e)}')
            try:
                cursor.close()
                conn.close()
            except:
                pass

    return StreamingResponse(
        generate(),
        media_type='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no'
        }
    )


def stream_reinit_database_with_step(step):
    def generate():
        yield sse_message('log', '🔄 开始重新初始化数据库...')
        
        # 清理 photos 文件夹
        try:
            import os
            import shutil
            from pathlib import Path
            
            photos_dir = Path(settings.UPLOAD_FOLDER)
            
            if photos_dir.exists():
                yield sse_message('log', '🗑️ 正在清理 photos 文件夹...')
                # 删除所有文件，保留目录
                for item in photos_dir.iterdir():
                    if item.is_file():
                        item.unlink()
                    elif item.is_dir():
                        shutil.rmtree(item)
                yield sse_message('info', '✅ photos 文件夹已清空')
            else:
                photos_dir.mkdir(parents=True, exist_ok=True)
                yield sse_message('info', '📁 创建 photos 文件夹')
        except Exception as e:
            yield sse_message('warning', f'⚠️ 清理 photos 文件夹失败: {str(e)}')
        
        conn = None
        cursor = None
        try:
            conn = mysql.connector.connect(
                host=db_config['host'],
                user=db_config['user'],
                password=db_config['password']
            )
            cursor = conn.cursor()
            db_name = db_config['database']
            yield sse_message('log', f'🗑️ 删除旧数据库 {db_name}...')
            cursor.execute(f"DROP DATABASE IF EXISTS `{db_name}`")
            yield sse_message('info', f'🆕 创建新数据库 {db_name}...')
            cursor.execute(f"""
                CREATE DATABASE `{db_name}` DEFAULT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
            """)
            cursor.execute(f"USE `{db_name}`")

            create_statements = CREATE_TABLES_SQL

            yield sse_message('log', '🔨 正在创建表结构（按外键依赖顺序）...')
            for i, stmt in enumerate(create_statements, 1):
                if not isinstance(stmt, str):
                    yield sse_message('error', f'❌ 非法建表语句：{repr(stmt)}')
                    continue
                try:
                    first_line = stmt.strip().split('\n')[0].strip()
                    parts = first_line.split()
                    table_name = parts[2] if len(parts) >= 3 and parts[0].upper() == 'CREATE' and parts[1].upper() == 'TABLE' else 'unknown_table'
                except:
                    table_name = 'unknown_table'
                cursor.execute(stmt)
                yield sse_message('info', f'✅ 第 {i}/6 张表创建成功: {table_name}')
            conn.commit()

            # 保险修复：若 InspectionRecord 缺少 data_type 列，则自动补全并创建索引
            try:
                cursor.execute("SHOW COLUMNS FROM InspectionRecord LIKE 'data_type'")
                dt_col = cursor.fetchone()
                if not dt_col:
                    yield sse_message('warning', '⚠️ 检测到缺少 data_type 列，正在自动修复...')
                    cursor.execute("""
                        ALTER TABLE InspectionRecord 
                        ADD COLUMN data_type ENUM('real','test') DEFAULT 'real' COMMENT '数据类型：real=真实数据，test=测试数据'
                    """)
                    cursor.execute("CREATE INDEX IF NOT EXISTS idx_data_type ON InspectionRecord(data_type)")
                    conn.commit()
                    yield sse_message('success', '✅ 已自动添加 data_type 列与索引')
            except Exception as e:
                try:
                    # 兼容低版本 MySQL 不支持 IF NOT EXISTS 语法时，容错处理索引创建
                    cursor.execute("SHOW INDEX FROM InspectionRecord WHERE Key_name='idx_data_type'")
                    has_idx = cursor.fetchone()
                    if not has_idx:
                        cursor.execute("CREATE INDEX idx_data_type ON InspectionRecord(data_type)")
                        conn.commit()
                        yield sse_message('info', '🧩 已补充创建 idx_data_type 索引')
                except Exception:
                    pass

            if step == '1':
                yield sse_message('success', '✅ 数据库表结构初始化完成！')
            elif step == 'all':
                yield sse_message('log', '📝 正在插入测试数据...')
                try:
                    TEST_DATA = get_test_data()  # 在使用时获取测试数据
                    for table_name, rows in TEST_DATA.items():
                        if not rows: continue
                        for row in rows:
                            cols = []
                            vals = []
                            sql_vals = []
                            for k, v in row.items():
                                cols.append(k)
                                if v == "NOW()":
                                    sql_vals.append("NOW()")
                                else:
                                    sql_vals.append("%s")
                                    vals.append(v)
                            insert_sql = f"INSERT INTO {table_name} ({', '.join(cols)}) VALUES ({', '.join(sql_vals)})"
                            cursor.execute(insert_sql, vals)
                            yield sse_message('debug', f' ➕ 插入 {table_name}: {vals[:3]}...')
                    yield sse_message('info', '📊 测试数据插入完成')
                    cursor.execute("SELECT COUNT(*) FROM InspectionRecord")
                    count = cursor.fetchone()[0]
                    yield sse_message('info', f'🔍 巡查记录总数: {count}')
                    conn.commit()
                    yield sse_message('success', '✅ 数据库完整重置成功（含测试数据）！')
                except Exception as e:
                    yield sse_message('error', f'❌ 插入测试数据失败: {str(e)}')
                    conn.rollback()
                    raise
            else:
                yield sse_message('error', f'❌ 未知的重置模式: {step}')
                return
        except Exception as e:
            yield sse_message('error', f'❌ 初始化失败: {str(e)}')
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
            yield sse_message('complete', '🏁 重置操作结束')
    return StreamingResponse(generate(), media_type='text/event-stream')


def save_photo_to_record(record_id: int, photo_type: str, file_path: str, file_name: str, file_size: int):
    """
    保存照片记录到数据库
    
    Args:
        record_id: 巡查记录ID
        photo_type: 照片类型（'upload', 'after_fix'）
        file_path: 文件路径
        file_name: 文件名
        file_size: 文件大小（字节）
    
    Returns:
        照片ID或None
    """
    conn = None
    try:
        conn = mysql.connector.connect(
            host=db_config['host'],
            user=db_config['user'],
            password=db_config['password'],
            database=db_config['database'],
            charset='utf8mb4'
        )
        cursor = conn.cursor()
        
        cursor.execute(
            """
            INSERT INTO Photo (record_id, photo_type, file_path, file_name, file_size, is_watermarked)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (record_id, photo_type, file_path, file_name, file_size, 1)
        )
        conn.commit()
        photo_id = cursor.lastrowid
        cursor.close()
        conn.close()
        
        return photo_id
    except Exception as e:
        print(f"保存照片失败: {e}")
        if conn:
            try:
                conn.rollback()
            except:
                pass
        return None


def clean_test_data():
    """
    删除所有测试数据（data_type='test'的记录及关联照片）
    
    Returns:
        dict: {'success': bool, 'deleted_count': int, 'photos_deleted': int}
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # 1. 查询测试数据的record_id
        cursor.execute("SELECT record_id FROM InspectionRecord WHERE data_type = 'test'")
        test_record_ids = [row[0] for row in cursor.fetchall()]
        
        if not test_record_ids:
            return {'success': True, 'deleted_count': 0, 'photos_deleted': 0}
        
        # 2. 删除关联照片文件和数据库记录
        photos_dir = settings.UPLOAD_FOLDER
        photos_deleted = 0
        
        placeholders = ','.join(['%s'] * len(test_record_ids))
        cursor.execute(f"SELECT file_path FROM Photo WHERE record_id IN ({placeholders})", test_record_ids)
        photo_paths = [row[0] for row in cursor.fetchall()]
        
        for photo_path in photo_paths:
            try:
                if os.path.exists(photo_path):
                    os.remove(photo_path)
                    photos_deleted += 1
            except Exception as e:
                print(f"删除照片失败 {photo_path}: {e}")
        
        # 3. 删除Photo表记录
        cursor.execute(f"DELETE FROM Photo WHERE record_id IN ({placeholders})", test_record_ids)
        
        # 4. 删除InspectionRecord表记录
        cursor.execute(f"DELETE FROM InspectionRecord WHERE data_type = 'test'")
        deleted_count = cursor.rowcount
        
        conn.commit()
        return {
            'success': True,
            'deleted_count': deleted_count,
            'photos_deleted': photos_deleted
        }
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cursor.close()
        conn.close()